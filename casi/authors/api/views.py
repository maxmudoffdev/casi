import io

from django.shortcuts import get_object_or_404
from docutils.nodes import authors
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework import  status
from casi.authors.api.filters import AuthorFilter
from casi.authors.api.pagination import AuthorPagination
from casi.authors.api.serializers import AuthorSerializersPrivate
from casi.authors.models import Author
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class AllAuthorView(APIView):
    permission_classes = [AllowAny]
    pagination_class = AuthorPagination
    def get(self,request):
        queryset = Author.objects.all()
        filterset = AuthorFilter(request.GET, queryset=queryset)
        if filterset.is_valid():
            queryset = filterset.qs
        paginator = AuthorPagination()

        ordering = request.query_params.get("ordering","first_name")
        allowed = ["last_name", "first_name","email", "created_at", "-last_name", "-first_name", "-created_at"]
        if ordering in allowed:
            queryset = queryset.order_by(ordering)

        page = paginator.paginate_queryset(queryset,request)
        serializers = AuthorSerializersPrivate(page,many=True)


        return paginator.get_paginated_response(serializers.data)

    def post(self,request):
        many = isinstance(request.data,list)
        serializers = AuthorSerializersPrivate(data=request.data,many=True)
        serializers.is_valid(raise_exception=True)
        if many:
            authors = [Author(**item) for item in serializers.validated_data]
            Author.objects.bulk_create(authors)

        else:
            serializers.save()
        res_data = {
                "message":"Author is created",
                "data":serializers.data
            }
        return Response(data=res_data,status=status.HTTP_201_CREATED)


class AuthorDetailView(APIView):
    permission_classes = [AllowAny]
    def get_author(self,id):
        return get_object_or_404(Author,id=id)
    def get(self,request,id):
        serializers = AuthorSerializersPrivate(self.get_author(id))
        return Response(data=serializers.data,status=status.HTTP_200_OK)

    def put(self, request, id):
        serializers = AuthorSerializersPrivate(self.get_author(id),data=request.data)
        serializers.is_valid(raise_exception=True)
        serializers.save()
        res_data = {
                "message": "Author is updated",
                "data": serializers.data
            }
        return Response(data=res_data,status=status.HTTP_200_OK)



    def patch(self, request, id):
        serializers = AuthorSerializersPrivate(self.get_author(id), data=request.data,partial=True)
        serializers.is_valid(raise_exception=True)
        serializers.save()
        res_data = {
                "message": "Author is updated",
                "data": serializers.data
            }
        return Response(data=res_data, status=status.HTTP_200_OK)

    def delete(self, request, id):
        author = self.get_author(id)
        author.is_active = False
        author.save()
        return Response(status=status.HTTP_204_NO_CONTENT)

class DowloadAllUserView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        queryset = Author.objects.all()

        # AuthorFilter — qayta ishlatamiz ✅
        filterset = AuthorFilter(request.GET, queryset=queryset)

        if filterset.is_valid():
            queryset = filterset.qs
        ordering = request.query_params.get("ordering", "first_name")
        allowed = ["last_name", "first_name", "created_at", "-last_name", "-first_name", "-created_at"]
        if ordering in allowed:
            queryset = queryset.order_by(ordering)

        authors = queryset.values(
            "first_name", "last_name", "affiliation", "email", "orc_id"
        )

        df = pd.DataFrame(list(authors))
        df.columns = ["First Name", "Last Name", "Affiliation", "Email", "ORCID"]
        df["ORCID"] = df["ORCID"].fillna("")  # null → bo'sh

        buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Authors")

            # Stylelash
            ws = writer.sheets["Authors"]

            # Header style
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(fill_type="solid", fgColor="2E75B6")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for cell in ws[1]:  # 1-qator — header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            # Qatorlar
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.border = border

            # Ustun kengligi
            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 40
            ws.column_dimensions["D"].width = 30
            ws.column_dimensions["E"].width = 25

        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="authors.xlsx"'
        return response

